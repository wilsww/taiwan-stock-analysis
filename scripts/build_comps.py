"""
記憶體產業同業比較分析 — build_comps.py
Memory Sector Comparable Company Analysis
Generated: 2026-03-22
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ── Colour palette ──────────────────────────────────────────
DARK_NAVY  = "17365D"
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E4057"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
BLUE_IN    = "2E75B6"   # hardcoded-input text
GREY_NA    = "999999"   # N/A text
AMBER_WARN = "C55A11"   # orange for warnings
GREEN_OK   = "375623"   # green for discount
YELLOW_BG  = "FFF2CC"
GREEN_BG   = "EBF3E8"
BLUE_BG    = "EEF2F7"

# ── Style helpers ────────────────────────────────────────────
def _font(cell, bold=False, color=None, size=10, italic=False):
    color = color or "000000"
    cell.font = Font(name="Times New Roman", bold=bold, color=color,
                     size=size, italic=italic)

def _fill(cell, color=WHITE):
    cell.fill = PatternFill("solid", fgColor=color)

def _align(cell, h="center", v="center", wrap=False):
    cell.alignment = Alignment(horizontal=h, vertical=v,
                                wrap_text=wrap, shrink_to_fit=False)

def hdr(cell, text, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    cell.value = text
    _font(cell, bold=bold, color=fg, size=size)
    _fill(cell, bg)
    _align(cell)

def col_hdr(cell, text):
    cell.value = text
    _font(cell, bold=True, color="000000", size=9)
    _fill(cell, LIGHT_BLUE)
    _align(cell, wrap=True)

def inp(cell, value, fmt=None):
    """Hardcoded input — blue text"""
    cell.value = value
    _font(cell, color=BLUE_IN, size=10)
    _fill(cell, WHITE)
    _align(cell)
    if fmt:
        cell.number_format = fmt

def frm(cell, formula, fmt=None):
    """Formula — black text"""
    cell.value = formula
    _font(cell, color="000000", size=10)
    _fill(cell, WHITE)
    _align(cell)
    if fmt:
        cell.number_format = fmt

def na(cell, text="N/A"):
    """Not available"""
    cell.value = text
    _font(cell, color=GREY_NA, size=10, italic=True)
    _fill(cell, WHITE)
    _align(cell)

def stat(cell, val, fmt=None, bold=False):
    cell.value = val
    _font(cell, color="000000", size=10, bold=bold)
    _fill(cell, LIGHT_GREY)
    _align(cell)
    if fmt:
        cell.number_format = fmt

def lbl(cell, text, bold=False):
    cell.value = text
    _font(cell, bold=bold, color="000000", size=10)
    _fill(cell, LIGHT_GREY)
    _align(cell, h="left")

def note_cell(cell, text):
    cell.value = text
    _font(cell, color="000000", size=9)
    _fill(cell, WHITE)
    _align(cell, h="left", wrap=True)

def add_comment(cell, text, author="Comps Model"):
    c = Comment(text, author)
    c.width  = 280
    c.height = 80
    cell.comment = c

# ── Workbook setup ───────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "記憶體同業比較"
ws.sheet_view.showGridLines = False

# ── Row / col dimensions ─────────────────────────────────────
col_widths = [22, 12, 7, 11, 11, 11, 10, 12, 14, 11, 13, 12, 50]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ── Helper: fill entire row with a bg color ──────────────────
def fill_row(row, col_start, col_end, bg):
    for c in range(col_start, col_end + 1):
        _fill(ws.cell(row=row, column=c), bg)

# =============================================================
# HEADER BLOCK  (Rows 1–3)
# =============================================================
ws.merge_cells("A1:M1")
hdr(ws["A1"],
    "記憶體產業同業比較分析   MEMORY SECTOR COMPARABLE COMPANY ANALYSIS",
    bg=DARK_NAVY, size=13)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:M2")
hdr(ws["A2"],
    "華邦電子 (2344.TW)  •  南亞科技 (2408.TW)  •  Micron Technology (MU)  •  SanDisk Corporation (SNDK)",
    bg=DARK_BLUE, size=10)
ws.row_dimensions[2].height = 20

ws.merge_cells("A3:M3")
hdr(ws["A3"],
    "截至 2026年3月22日  |  台股：新台幣億元 (NT$100M)  |  美股：美元億元 (US$100M)  |  參考匯率 32 NT$/USD",
    bg=MID_BLUE, size=9)
ws.row_dimensions[3].height = 16

ws.row_dimensions[4].height = 6   # spacer

# =============================================================
# SECTION 1 — OPERATING STATISTICS  (Rows 5–16)
# =============================================================
ws.merge_cells("A5:M5")
hdr(ws["A5"],
    "▌ 第一部分：營運指標與財務績效   OPERATING STATISTICS & FINANCIAL METRICS   （基準：FY2025）",
    bg=DARK_NAVY, size=10)
ws.row_dimensions[5].height = 22

# Col headers  (row 6)
op_hdrs = [
    "公司 / Company", "代碼 Ticker", "幣別",
    "全年營收\n(億)", "YoY 成長%",
    "毛利率%\nGross Margin",
    "EBITDA\n(億)", "EBITDA%",
    "營業利益率%\nOp. Margin",
    "EPS\n(全年 2025)",
    "Q4 2025\n毛利率%",
    "每股淨值\nBV / Share",
    "備註 Notes",
]
for ci, h in enumerate(op_hdrs, 1):
    col_hdr(ws.cell(row=6, column=ci), h)
ws.row_dimensions[6].height = 44

# ── Operating data rows 7–10 ─────────────────────────────────
#  Col mapping:  A=1 Company  B=2 Ticker  C=3 Currency
#                D=4 Revenue  E=5 YoY%   F=6 GrossMargin
#                G=7 EBITDA   H=8 EBITDA%  I=9 OpMargin
#                J=10 EPS     K=11 Q4GM   L=12 BVShare  M=13 Notes

OP = [
    # Winbond  (row 7)
    dict(
        company="華邦電子  Winbond", ticker="2344.TW", cur="NT$",
        rev=894.06, yoy=0.0955,
        gm=None,    gm_note="~27%e (含新唐稀釋~9pp；純記憶體段落更高)",
        ebitda=None, ebitda_m=None,
        op_m=None,
        eps=0.88,
        q4gm=0.4186, q4gm_note="Q4合並毛利率41.86%；記憶體段落50.8%",
        bv=21.0,    bv_note="估計值；待Q4 2025年報確認",
        notes="含新唐(4919)子公司合並；純記憶體段落毛利率更高；全年EPS受Q1-Q2虧損拖累",
    ),
    # Nanya  (row 8) — official press release
    dict(
        company="南亞科技  Nanya Tech", ticker="2408.TW", cur="NT$",
        rev=665.87, yoy=0.951,
        gm=0.225,  gm_note="官方財報；Q4毛利率49.0%，Q3=18.5%，Q2=-20.6%",
        ebitda=177.0, ebitda_m=0.266,
        op_m=0.079,
        eps=2.13,
        q4gm=0.490, q4gm_note="官方Q4 2025自結新聞稿；Q4單季EPS 3.58",
        bv=54.99,  bv_note="官方數據；截至2025/12/31",
        notes="官方Q4 2025自結財報(2026/01/19)；全年轉虧為盈；Q4獲利NT$110.83億",
    ),
    # Micron  (row 9) — FY2025 ending Aug 2025
    dict(
        company="Micron Technology", ticker="MU", cur="US$",
        rev=374.0, yoy=0.492,
        gm=0.398,  gm_note="FY2025 10-K; DRAM prices +mid-60% YoY",
        ebitda=181.0, ebitda_m=0.484,
        op_m=None,
        eps=7.59,
        q4gm=None, q4gm_note="",
        bv=52.23,  bv_note="Nov 2025 quarter; TTM EPS已達$21.26",
        notes="FY2025截至2025年8月；TTM Revenue $581億，TTM EPS $21.26；Q2 FY2026 guided $18.7B revenue",
    ),
    # SanDisk  (row 10) — FY2025 ending Jun 2025
    dict(
        company="SanDisk Corporation", ticker="SNDK", cur="US$",
        rev=73.6, yoy=0.104,
        gm=0.301,  gm_note="FY2025 ending Jun 2025; LTM gross margin 32.4%",
        ebitda=-12.1, ebitda_m=-0.164,
        op_m=None,
        eps=-11.32,
        q4gm=0.509, q4gm_note="Q2 FY2026 (most recent quarter, ended ~Jan 2026)",
        bv=68.66,  bv_note="As of Jan 2026 quarter-end; total equity ~$9.96B",
        notes="FY2025含大額減損(-$1.93B Q3 FY2025)；LTM Revenue $89.3億；Q2 FY2026 EPS $5.15顯示快速復甦",
    ),
]

for ri, d in enumerate(OP):
    r = 7 + ri
    ws.row_dimensions[r].height = 24

    # A: company
    c = ws.cell(r, 1, d["company"])
    _font(c, bold=True, size=10)
    _fill(c, WHITE)
    _align(c, h="left")

    inp(ws.cell(r, 2), d["ticker"])
    inp(ws.cell(r, 3), d["cur"])

    inp(ws.cell(r, 4), d["rev"], "#,##0.00")
    add_comment(ws.cell(r, 4),
        f"Source: {d.get('notes','')[:80]}")

    inp(ws.cell(r, 5), d["yoy"], "0.0%")

    if d["gm"] is not None:
        inp(ws.cell(r, 6), d["gm"], "0.0%")
        add_comment(ws.cell(r, 6), d["gm_note"])
    else:
        c = ws.cell(r, 6, "~27%e")
        _font(c, color=BLUE_IN, size=10, italic=True)
        _fill(c, WHITE); _align(c)
        add_comment(ws.cell(r, 6), d["gm_note"])

    if d["ebitda"] is not None:
        inp(ws.cell(r, 7), d["ebitda"], "#,##0.0")
    else:
        na(ws.cell(r, 7), "N/Ae")

    if d["ebitda_m"] is not None:
        inp(ws.cell(r, 8), d["ebitda_m"], "0.0%")
    else:
        na(ws.cell(r, 8), "N/Ae")

    if d["op_m"] is not None:
        inp(ws.cell(r, 9), d["op_m"], "0.0%")
    else:
        na(ws.cell(r, 9), "N/A")

    inp(ws.cell(r, 10), d["eps"], "0.00")

    if d["q4gm"] is not None:
        inp(ws.cell(r, 11), d["q4gm"], "0.0%")
        if d["q4gm_note"]:
            add_comment(ws.cell(r, 11), d["q4gm_note"])
    else:
        na(ws.cell(r, 11))

    inp(ws.cell(r, 12), d["bv"])
    add_comment(ws.cell(r, 12), d["bv_note"])

    note_cell(ws.cell(r, 13), d["notes"])

# ── Blank row 11 ─────────────────────────────────────────────
ws.row_dimensions[11].height = 6
fill_row(11, 1, 13, LIGHT_GREY)

# ── Stats rows 12–16 ─────────────────────────────────────────
STAT_LABELS = [
    ("最大值 Maximum",      False),
    ("75th 百分位 75th Pct.", False),
    ("中位數 Median",        True),
    ("25th 百分位 25th Pct.", False),
    ("最小值 Minimum",       False),
]
STAT_FUNCS = ["MAX", "QUARTILE({r},3)", "MEDIAN", "QUARTILE({r},1)", "MIN"]

for si, ((slbl, sbold), sfunc) in enumerate(zip(STAT_LABELS, STAT_FUNCS)):
    sr = 12 + si
    ws.row_dimensions[sr].height = 18

    ws.merge_cells(f"A{sr}:C{sr}")
    lbl(ws[f"A{sr}"], slbl, bold=sbold)

    # D: revenue — not comparable across currencies
    stat(ws.cell(sr, 4), "—"); ws.cell(sr,4).number_format = "@"

    def mk_stat(col_letter, rng, fmt, bold=False):
        r_ref = rng
        if sfunc.startswith("QUARTILE"):
            q = sfunc.split(",")[1].rstrip(")")
            formula = f"=QUARTILE({r_ref},{q})"
        else:
            formula = f"={sfunc.split('(')[0]}({r_ref})"
        stat(ws.cell(sr, ord(col_letter) - 64), formula, fmt, bold=bold and sbold)

    # E: YoY% — all 4 rows
    mk_stat("E", "E7:E10", "0.0%", bold=True)

    # F: Gross Margin — rows 8-10 only (Winbond is text estimate)
    mk_stat("F", "F8:F10", "0.0%", bold=True)

    # G, H, I: EBITDA / margins — only rows 8-9 have numeric data
    for ci_let, rng in [("G","G8:G9"),("H","H8:H9"),("I","I8:I8")]:
        mk_stat(ci_let, rng, "0.0")

    # J: EPS — all 4
    mk_stat("J", "J7:J10", "0.00", bold=True)

    # K: Q4 Gross Margin — rows 7,8,10 (MU is N/A)
    if sfunc.startswith("QUARTILE"):
        q = sfunc.split(",")[1].rstrip(")")
        fk = f"=QUARTILE((K7,K8,K10),{q})"
    elif sfunc.startswith("MAX"):
        fk = "=MAX(K7,K8,K10)"
    elif sfunc.startswith("MEDIAN"):
        fk = "=MEDIAN(K7,K8,K10)"
    else:
        fk = "=MIN(K7,K8,K10)"
    stat(ws.cell(sr, 11), fk, "0.0%", bold=sbold)

    # L, M: blank
    for ci in [12, 13]:
        stat(ws.cell(sr, ci), "")

# =============================================================
# SECTION 2 — VALUATION MULTIPLES  (Rows 18–29)
# =============================================================
ws.row_dimensions[17].height = 8

ws.merge_cells("A18:M18")
hdr(ws["A18"],
    "▌ 第二部分：估值倍數比較   VALUATION MULTIPLES   （截至 2026年3月20日收盤）",
    bg=DARK_NAVY, size=10)
ws.row_dimensions[18].height = 22

val_hdrs = [
    "公司 / Company", "代碼 Ticker", "幣別",
    "股價\nPrice",
    "市值\n(億)",
    "每股淨值\nBV/Share",
    "P / B",
    "P/E\nTrailing",
    "P/E Forward\n2026E",
    "EV (億)e",
    "EV/Revenue\n(FY25)",
    "EV/EBITDA\n(FY25)",
    "估值評等 vs Peers",
]
for ci, h in enumerate(val_hdrs, 1):
    col_hdr(ws.cell(19, ci), h)
ws.row_dimensions[19].height = 44

# ── Valuation data rows 20–23 ────────────────────────────────
# For formulas that cross-reference Section 1, we reference rows 7-10
# Row offsets: Winbond=r7, Nanya=r8, Micron=r9, SanDisk=r10

VAL = [
    # Winbond (row 20) — refs op row 7
    dict(
        company="華邦電子  Winbond", ticker="2344.TW", cur="NT$",
        price=110, mktcap=4950,
        bv_ref="L7",          # BV/share is in L7 (=21, estimate)
        pe_trail_ref="J7",    # EPS in J7
        pe_fwd=23.9, pe_fwd_note="Bloomberg consensus 2026E EPS NT$4.6；JPMorgan NT$6.3",
        ev=None, ev_rev=None, ev_ebitda=None,
        verdict="⚠ 相對溢價", verdict_bg=AMBER_WARN,
        detail="PB 5.24x > 南亞科4.25x；Forward PE 23.9x(4.6E)為四家最高；溢價缺乏EPS槓桿支撐",
    ),
    # Nanya (row 21) — refs op row 8
    dict(
        company="南亞科技  Nanya Tech", ticker="2408.TW", cur="NT$",
        price=234, mktcap=7251,
        bv_ref="L8",
        pe_trail_ref="J8",
        pe_fwd=14.1, pe_fwd_note="基於JP analysts 2026E EPS NT$16.54（保守）；凱基估NT$33.16",
        ev=None, ev_rev=None, ev_ebitda=None,
        verdict="✓ 合理/偏低", verdict_bg=GREEN_OK,
        detail="PB 4.25x；Forward PE 14.1x(保守)～7.1x(樂觀)；週期EPS槓桿最大；為四家中估值最具吸引力",
    ),
    # Micron (row 22) — refs op row 9
    dict(
        company="Micron Technology", ticker="MU", cur="US$",
        price=422.9, mktcap=4769,
        bv_ref="L9",
        pe_trail_ref=None, pe_trail_val=19.89,
        pe_fwd=4.72, pe_fwd_note="stockanalysis.com consensus; TTM EPS $21.26 已遠超FY25的$7.59",
        ev=4819, ev_rev_formula="=J22/D9", ev_ebitda_formula="=J22/G9",
        verdict="✓ 折扣", verdict_bg=GREEN_OK,
        detail="Forward PE 4.72x為四家最便宜；PB 8.1x但TTM EPS成長強勁；市場定價偏保守",
    ),
    # SanDisk (row 23) — refs op row 10
    dict(
        company="SanDisk Corporation", ticker="SNDK", cur="US$",
        price=709.71, mktcap=1047,
        bv_ref="L10",
        pe_trail_ref=None, pe_trail_val=None,
        pe_fwd=9.05, pe_fwd_note="Forward PE 9.05x; Q2 FY2026 EPS $5.15；全年2026供給幾乎售罄",
        ev=1041, ev_rev_formula="=J23/D10", ev_ebitda_formula=None,
        verdict="⚠ 高溢價", verdict_bg=AMBER_WARN,
        detail="PB 10.3x最高；FY25 EPS負但市場定價NAND超級循環；EV/Revenue 14.1x(FY25)偏高",
    ),
]

for ri, d in enumerate(VAL):
    r = 20 + ri
    ws.row_dimensions[r].height = 26

    c = ws.cell(r, 1, d["company"])
    _font(c, bold=True, size=10); _fill(c, WHITE); _align(c, h="left")

    inp(ws.cell(r, 2), d["ticker"])
    inp(ws.cell(r, 3), d["cur"])
    inp(ws.cell(r, 4), d["price"], "#,##0.00")
    inp(ws.cell(r, 5), d["mktcap"], "#,##0")

    # BV/share  (cross-ref to Section 1 BV column L)
    op_row = 7 + ri
    frm(ws.cell(r, 6), f"=L{op_row}", "#,##0.00")

    # P/B
    frm(ws.cell(r, 7), f"=D{r}/F{r}", '0.00"x"')

    # P/E trailing
    if d.get("pe_trail_ref"):
        frm(ws.cell(r, 8), f"=D{r}/{d['pe_trail_ref']}", '0.0"x"')
    elif d.get("pe_trail_val") is not None:
        inp(ws.cell(r, 8), d["pe_trail_val"], '0.0"x"')
        add_comment(ws.cell(r, 8), "Source: stockanalysis.com, TTM basis")
    else:
        na(ws.cell(r, 8), "N/M")

    # P/E forward
    inp(ws.cell(r, 9), d["pe_fwd"], '0.0"x"')
    add_comment(ws.cell(r, 9), d["pe_fwd_note"])

    # EV
    if d.get("ev") is not None:
        inp(ws.cell(r, 10), d["ev"], "#,##0")
        add_comment(ws.cell(r, 10), "EV = Market Cap + Net Debt (approximate)")
    else:
        na(ws.cell(r, 10), "N/Ae")

    # EV/Revenue
    if d.get("ev_rev_formula"):
        frm(ws.cell(r, 11), d["ev_rev_formula"], '0.0"x"')
    else:
        na(ws.cell(r, 11), "N/Ae")

    # EV/EBITDA
    if d.get("ev_ebitda_formula"):
        frm(ws.cell(r, 12), d["ev_ebitda_formula"], '0.0"x"')
    else:
        na(ws.cell(r, 12), "N/M" if ri == 3 else "N/Ae")

    # Assessment
    c_v = ws.cell(r, 13, d["verdict"])
    _font(c_v, bold=True, color=WHITE, size=9)
    _fill(c_v, d["verdict_bg"])
    _align(c_v, h="center", wrap=True)

# ── Blank row 24 ─────────────────────────────────────────────
ws.row_dimensions[24].height = 6
fill_row(24, 1, 13, LIGHT_GREY)

# ── Val stats rows 25–29 ─────────────────────────────────────
for si, ((slbl, sbold), sfunc) in enumerate(zip(STAT_LABELS, STAT_FUNCS)):
    sr = 25 + si
    ws.row_dimensions[sr].height = 18

    ws.merge_cells(f"A{sr}:F{sr}")
    lbl(ws[f"A{sr}"], slbl, bold=sbold)

    # P/B (col 7)
    mk_stat("G", "G20:G23", '0.00"x"', bold=True)

    # P/E trailing — only rows 20-22 (SNDK N/M)
    if sfunc.startswith("QUARTILE"):
        q = sfunc.split(",")[1].rstrip(")")
        ft = f"=QUARTILE(H20:H22,{q})"
    elif sfunc.startswith("MAX"):
        ft = "=MAX(H20:H22)"
    elif sfunc.startswith("MEDIAN"):
        ft = "=MEDIAN(H20:H22)"
    else:
        ft = "=MIN(H20:H22)"
    stat(ws.cell(sr, 8), ft, '0.0"x"', bold=sbold)

    # P/E forward (col 9) — all 4
    mk_stat("I", "I20:I23", '0.0"x"', bold=True)

    # EV, EV/Rev, EV/EBITDA — only US stocks rows 22-23
    for ci_let, rng, fmt in [("J","J22:J23","#,##0"),
                              ("K","K22:K23",'0.0"x"')]:
        if sfunc.startswith("QUARTILE"):
            q = sfunc.split(",")[1].rstrip(")")
            fv = f"=QUARTILE({rng},{q})"
        elif sfunc.startswith("MAX"):
            fv = f"=MAX({rng})"
        elif sfunc.startswith("MEDIAN"):
            fv = f"=MEDIAN({rng})"
        else:
            fv = f"=MIN({rng})"
        stat(ws.cell(sr, ord(ci_let)-64), fv, fmt, bold=sbold)

    stat(ws.cell(sr, 12), "—")
    stat(ws.cell(sr, 13), "")

# =============================================================
# SECTION 3 — CYCLE POSITION & KEY INSIGHTS  (Rows 31–42)
# =============================================================
ws.row_dimensions[30].height = 8

ws.merge_cells("A31:M31")
hdr(ws["A31"],
    "▌ 第三部分：週期定位與核心結論   CYCLE POSITION & KEY INVESTMENT CONCLUSIONS",
    bg=DARK_NAVY, size=10)
ws.row_dimensions[31].height = 22

INSIGHTS = [
    ("① 華邦電 vs 南亞科  \n（最直接同業）",
     "⚠ PB 溢價 ~23%",
     AMBER_WARN, YELLOW_BG,
     "華邦電 P/B ~5.24x vs 南亞科 4.25x，溢價約23%。Forward PE 23.9x(4.6E) vs 南亞科 14.1x(16.54E)。"
     "南亞科週期 EPS 槓桿 8-18x，製程升級 1C→1D 節奏明確，Q4 毛利率 49%，"
     "股價仍有向上空間。華邦電溢價在財務上難以支撐，估值偏貴。"),

    ("② 華邦電 vs Micron  \n（全球龍頭）",
     "Forward PE 高出 5x",
     AMBER_WARN, YELLOW_BG,
     "MU Forward PE 4.72x（最便宜），TTM EPS $21.26 已大幅超越 FY25 的 $7.59；"
     "華邦電 Forward PE 23.9x。即使考量規模與技術差距，MU 的週期性折扣更深，"
     "顯示市場對台灣 Specialty DRAM 給予相對流動性/成長溢價，但絕對估值偏高。"),

    ("③ 華邦電 vs SanDisk  \n（NAND 超循環）",
     "PB 折扣 (NAND>DRAM)",
     GREEN_OK, GREEN_BG,
     "SNDK PB 10.3x 顯示市場對 NAND 超循環的高度定價（2026 年 NAND 供給幾乎全年售罄）。"
     "華邦電 PB 5.24x 相比 NAND 週期股明顯折扣，反映 Specialty DRAM 週期力道不及 NAND。"
     "NAND 復甦週期較 DRAM 超前，SNDK 溢價有一定合理性，但 10x PB 也偏激進。"),

    ("④ 關鍵投資結論",
     "四家中估值最貴",
     AMBER_WARN, YELLOW_BG,
     "華邦電在四家中 Forward PE 最高、P/B 溢價最大，不具備最高估值的財務支撐。"
     "若使用純記憶體段落拆分（排除新唐稀釋），估值可能更合理，但需等 Q4 2025 正式財報。"
     "目前股價 NT$110 接近部分外資目標價上限（NT$83-120）。若 2026E EPS 兌現 NT$6-7，"
     "PE 將降至 15-18x，估值合理化需靠獲利成長而非股價上漲。"),

    ("⑤ 四大常駐風險",
     "必須定期更新",
     MID_BLUE, BLUE_BG,
     "① CXMT：DDR3/DDR4 市場價格壓力持續，但 EUV 限制使進階製程良率落後。"
     "② 三星 P4 HBM4：2027H1 供給衝擊視窗，需每季確認。"
     "③ AI CapEx 成長率：任一大廠 YoY 降幅 >10% 為警戒訊號。"
     "④ 地緣政治：InP 基板出口管制已解除(2025Q3)，台海情勢持續監控。"),
]

for ri, (title, verdict, vbg, dbg, detail) in enumerate(INSIGHTS):
    r = 32 + ri * 2
    ws.row_dimensions[r].height = 36
    ws.row_dimensions[r+1].height = 4

    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(r, 1, title)
    _font(c, bold=True, color=WHITE, size=9)
    _fill(c, DARK_BLUE); _align(c, h="left", wrap=True)

    ws.merge_cells(f"D{r}:E{r}")
    c = ws.cell(r, 4, verdict)
    _font(c, bold=True, color=WHITE, size=9)
    _fill(c, vbg); _align(c, h="center", wrap=True)

    ws.merge_cells(f"F{r}:M{r}")
    c = ws.cell(r, 6, detail)
    _font(c, color="000000", size=9)
    _fill(c, dbg); _align(c, h="left", wrap=True)

    # spacer row
    fill_row(r+1, 1, 13, WHITE)

# =============================================================
# SECTION 4 — DATA SOURCES & NOTES  (Rows 43+)
# =============================================================
ws.row_dimensions[42].height = 8

ws.merge_cells("A43:M43")
hdr(ws["A43"],
    "▌ 第四部分：資料來源與方法論   DATA SOURCES & METHODOLOGY",
    bg=DARK_NAVY, size=10)
ws.row_dimensions[43].height = 22

NOTES = [
    ("資料來源",
     "南亞科: 官方Q4 2025自結新聞稿(2026/01/19) ✓  |  華邦電: 法說會PDF + Web search  |  "
     "Micron: 10-K FY2025 & Form 10-Q  |  SanDisk: FY2025年報 & FY2026 Q1-Q2財報  |  "
     "股價/市值: 2026/03/20收盤"),
    ("標記說明",
     "藍字 = 直接輸入數據  |  黑字 = 公式計算  |  'e' 後綴 = 估計值  |  "
     "N/A = 不適用  |  N/M = 無意義 (如負數倍數)  |  ~ = 概略值"),
    ("幣別說明",
     "台股以 NT$億 (NT$100M) 計；美股以 US$億 (US$100M) 計。"
     "P/B、P/E、EV倍數可跨幣別比較；Revenue、Market Cap絕對數字不可直接比較。"
     "參考匯率 32 NT$/USD，NT$894億 ≈ US$27.9億。"),
    ("重要估計值",
     "① 華邦電 BV/Share ~NT$21為估計（實際值待Q4 2025正式年報）"
     "② 華邦電 FY2025毛利率~27%為加權估計（Q4=41.9%, Q3=46.7%, Q2=22.7%）"
     "③ 南亞科 EBITDA ~NT$177億 = 營業淨利NT$52.4億 + 估算D&A ~NT$125億"
     "④ Micron BV/Share $52.23截至2025年11月"),
    ("分析限制",
     "① 台股 EV 計算因缺乏即時淨負債數據而省略（建議補充MOPS財報）"
     "② 跨幣別 EV 倍數比較存在匯率風險"
     "③ 華邦電純記憶體段落估值需另行拆分（新唐合併稀釋毛利率~9pp）"
     "④ Forward EPS為共識預估，實際差異可能顯著"),
]

for ri, (label, content) in enumerate(NOTES):
    r = 44 + ri
    ws.row_dimensions[r].height = 32
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, label)
    _font(c, bold=True, color=WHITE, size=9)
    _fill(c, DARK_BLUE); _align(c, h="left")
    ws.merge_cells(f"C{r}:M{r}")
    c = ws.cell(r, 3, content)
    _font(c, color="000000", size=9)
    _fill(c, LIGHT_GREY); _align(c, h="left", wrap=True)

# =============================================================
# SAVE
# =============================================================
output_path = "/Users/wayne/Desktop/Invest/DRAM/Memory_Sector_Comps_2026Q1.xlsx"
wb.save(output_path)
print(f"✓ Saved: {output_path}")
