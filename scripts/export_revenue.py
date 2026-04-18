#!/usr/bin/env python3
"""
月營收資料匯出工具
從 SQLite 資料庫匯出 CSV + Excel（月營收 / 季度彙整 / 最新快照）
用途：python3 scripts/export_revenue.py [--format csv|excel|both] [--summary]
"""

import sqlite3
import csv
import argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from typing import Optional, List, Dict, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ── 設定 ──────────────────────────────────────────────────────
DATA_DIR = Path(__file__).parent.parent / "data"
REVENUE_DIR = DATA_DIR / "revenue"
DB_PATH  = REVENUE_DIR / "revenue.db"

COMPANY_NAMES = {
    "2408": "南亞科技",   "2344": "華邦電子",   "3363": "上詮光纖",
    "2455": "全新光電",   "4979": "華星光通",   "3081": "聯亞光電",
    "3105": "穩懋半導體", "4919": "新唐科技",
    # CPO 擴充觀察名單
    "6442": "光聖",       "4977": "眾達-KY",    "3163": "波若威",
    "2345": "智邦",       "6223": "旺矽",
    # IC 載板 PCB 標的
    "3037": "欣興電子",   "8046": "南亞電路板", "3189": "景碩科技",
}

THEMES = {
    "2408": "DRAM/HBM 超級循環",   "2344": "Specialty DRAM+NOR Flash",
    "3363": "CPO L1 光通訊",        "2455": "InP 磊晶 L1",
    "4979": "CW Laser 封裝 L2",     "3081": "InP 磊晶 L1",
    "3105": "GaAs 晶圓代工 L2",     "4919": "嵌入式控制器",
    "6442": "CPO L3 光纖連接器",    "4977": "CPO L3 光模組",
    "3163": "CPO L3 特種光纖",      "2345": "CPO L4 交換器系統",
    "6223": "CPO 測試 Probe Card",
    "3037": "ABF 載板 全球龍頭",    "8046": "ABF+BT 載板",
    "3189": "ABF+BT+CoWoS 基板",
}

# Excel 色彩
COL_HEADER  = "1E3A5F"   # 深藍標題
COL_POS_HI  = "C8E6C9"   # YoY > 50% 淺綠
COL_POS_MD  = "E8F5E9"   # YoY > 0%  極淡綠
COL_NEG     = "FFCDD2"   # YoY < -20% 淺紅
COL_NEUTRAL = "F5F5F5"   # 隔行
COL_ACCENT  = "FFF9C4"   # 最新快照用


def get_all_data() -> List[Dict]:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"找不到資料庫：{DB_PATH}\n請先執行 fetch_revenue.py")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("""
        SELECT ticker, year, month, revenue_m, mom_pct, yoy_pct,
               cum_revenue, cum_yoy_pct, source, updated_at
        FROM monthly_revenue
        ORDER BY ticker, year DESC, month DESC
    """)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def compute_quarterly(rows: List[Dict]) -> List[Dict]:
    bucket: dict = defaultdict(lambda: {"revenue_m": 0.0, "count": 0})
    for r in rows:
        q = (r["month"] - 1) // 3 + 1
        key = (r["ticker"], r["year"], q)
        if r.get("revenue_m"):
            bucket[key]["revenue_m"] += r["revenue_m"]
            bucket[key]["count"] += 1

    results = []
    for (ticker, year, quarter), d in bucket.items():
        if d["count"] == 3:
            results.append({
                "ticker": ticker,
                "company": COMPANY_NAMES.get(ticker, ticker),
                "year": year, "quarter": quarter,
                "revenue_m": round(d["revenue_m"], 1),
                "revenue_b": round(d["revenue_m"] / 1000, 3),
            })
    results.sort(key=lambda x: (x["ticker"], x["year"], x["quarter"]))
    return results


def get_latest_per_ticker(rows: List[Dict]) -> Dict:
    seen: dict = {}
    for r in rows:
        if r["ticker"] not in seen:
            seen[r["ticker"]] = r
    return seen


# ── CSV 匯出 ──────────────────────────────────────────────────
def export_csv(rows: List[Dict], quarterly: List[Dict]) -> Tuple:
    monthly_path    = REVENUE_DIR / "revenue_monthly.csv"
    quarterly_path  = REVENUE_DIR / "revenue_quarterly.csv"

    with open(monthly_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["代碼","公司","主題","年","月","當月營收(億)","月增率%",
                    "年增率%","累計營收(億)","累計年增率%","資料來源","更新時間"])
        for r in rows:
            rev_b = round(r["revenue_m"] / 1000, 3) if r.get("revenue_m") else ""
            cum_b = round(r["cum_revenue"] / 1000, 3) if r.get("cum_revenue") else ""
            w.writerow([
                r["ticker"], COMPANY_NAMES.get(r["ticker"], r["ticker"]),
                THEMES.get(r["ticker"], ""), r["year"], r["month"],
                rev_b, r.get("mom_pct",""), r.get("yoy_pct",""),
                cum_b, r.get("cum_yoy_pct",""),
                r.get("source",""), r.get("updated_at",""),
            ])

    with open(quarterly_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["代碼","公司","年","季度","季營收(億)"])
        for r in quarterly:
            w.writerow([r["ticker"], r["company"], r["year"],
                        f"Q{r['quarter']}", r["revenue_b"]])

    print(f"  ✅ CSV → {monthly_path.name}")
    print(f"  ✅ CSV → {quarterly_path.name}")
    return monthly_path, quarterly_path


# ── Excel 匯出 ────────────────────────────────────────────────
def _header_row(ws, headers: list, fill_hex: str = COL_HEADER):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(color="FFFFFF", bold=True, name="Arial")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _yoy_fill(yoy):
    if yoy is None:
        return None
    if yoy > 100:
        c = COL_POS_HI
    elif yoy > 0:
        c = COL_POS_MD
    elif yoy < -20:
        c = COL_NEG
    else:
        return None
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def export_excel(rows: List[Dict], quarterly: List[Dict]) -> Optional[Path]:
    if not HAS_EXCEL:
        print("  ⚠️  未安裝 openpyxl，跳過 Excel 輸出 → 執行：pip3 install openpyxl")
        return None

    excel_path = REVENUE_DIR / "revenue_tracker.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1：月營收明細 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "月營收明細"
    h1 = ["代碼","公司","主題","年","月","當月營收(億)","月增率%","年增率%",
          "累計營收(億)","累計年增率%","資料來源"]
    _header_row(ws1, h1)
    for i, r in enumerate(rows, 2):
        rev_b = round(r["revenue_m"] / 1000, 3) if r.get("revenue_m") else None
        cum_b = round(r["cum_revenue"] / 1000, 3) if r.get("cum_revenue") else None
        vals  = [r["ticker"], COMPANY_NAMES.get(r["ticker"], r["ticker"]),
                 THEMES.get(r["ticker"],""), r["year"], r["month"],
                 rev_b, r.get("mom_pct"), r.get("yoy_pct"),
                 cum_b, r.get("cum_yoy_pct"), r.get("source","")]
        row_fill = _yoy_fill(r.get("yoy_pct"))
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            if row_fill:
                cell.fill = row_fill
            if col in (6, 7, 8, 9, 10) and isinstance(val, (int, float)):
                cell.number_format = "0.000"
    _set_col_widths(ws1, [8,12,22,6,5,14,10,10,14,14,12])
    ws1.freeze_panes = "A2"

    # ── Sheet 2：季度彙整 ────────────────────────────────────
    ws2 = wb.create_sheet("季度彙整")
    h2 = ["代碼","公司","年","季度","季營收(億)"]
    _header_row(ws2, h2)
    prev_ticker = None
    for i, r in enumerate(quarterly, 2):
        fill = None
        if r["ticker"] != prev_ticker:
            prev_ticker = r["ticker"]
        vals = [r["ticker"], r["company"], r["year"], f"Q{r['quarter']}", r["revenue_b"]]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            if fill:
                cell.fill = fill
            if col == 5 and isinstance(val, (int, float)):
                cell.number_format = "0.000"
    _set_col_widths(ws2, [8, 12, 6, 8, 14])
    ws2.freeze_panes = "A2"

    # ── Sheet 3：最新快照 ────────────────────────────────────
    ws3 = wb.create_sheet("最新快照")
    accent_fill = PatternFill(start_color=COL_ACCENT, end_color=COL_ACCENT, fill_type="solid")
    h3 = ["代碼","公司","主題","最新月份","當月營收(億)","月增率%","年增率%","累計(億)","資料更新時間"]
    _header_row(ws3, h3)
    latest = get_latest_per_ticker(rows)
    ticker_order = ["2408","2344","3363","2455","4979","3081","3105","4919"]
    row_i = 2
    for ticker in ticker_order:
        r = latest.get(ticker)
        if not r:
            continue
        rev_b = round(r["revenue_m"] / 1000, 3) if r.get("revenue_m") else None
        cum_b = round(r["cum_revenue"] / 1000, 3) if r.get("cum_revenue") else None
        vals = [
            ticker, COMPANY_NAMES.get(ticker, ticker), THEMES.get(ticker,""),
            f"{r['year']}年{r['month']:02d}月",
            rev_b, r.get("mom_pct"), r.get("yoy_pct"), cum_b,
            r.get("updated_at","")[:16] if r.get("updated_at") else "",
        ]
        row_fill = _yoy_fill(r.get("yoy_pct")) or accent_fill
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_i, column=col, value=val)
            cell.fill = row_fill
            if col in (5, 6, 7, 8) and isinstance(val, (int, float)):
                cell.number_format = "0.000"
        row_i += 1
    _set_col_widths(ws3, [8, 12, 22, 12, 14, 10, 10, 12, 18])
    ws3.freeze_panes = "A2"

    wb.save(excel_path)
    print(f"  ✅ Excel → {excel_path.name}（3 工作表：月營收明細 / 季度彙整 / 最新快照）")
    return excel_path


# ── 文字快照（供 command 報告用）────────────────────────────
def print_snapshot(rows: List[Dict]):
    latest = get_latest_per_ticker(rows)
    ticker_order = ["2408","2344","3363","2455","4979","3081","3105","4919"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    print()
    print("╔" + "═" * 76 + "╗")
    print(f"║  台股月營收快照  |  更新時間：{now:<45}║")
    print("╠" + "═" * 76 + "╣")
    print(f"║  {'代碼':<5} {'公司':<8} {'最新月份':<10} {'當月營收(億)':>12} {'MoM%':>8} {'YoY%':>10}  ║")
    print("╠" + "─" * 76 + "╣")
    for ticker in ticker_order:
        r = latest.get(ticker)
        if not r:
            print(f"║  {ticker:<5} {COMPANY_NAMES.get(ticker,ticker):<8} {'— 無資料 —':<10}{'':<35}║")
            continue
        rev_b  = r["revenue_m"] / 1000 if r.get("revenue_m") else 0
        mom_s  = f"{r['mom_pct']:+.1f}%" if r.get("mom_pct") is not None else "  N/A"
        yoy_s  = f"{r['yoy_pct']:+.1f}%" if r.get("yoy_pct") is not None else "  N/A"
        period = f"{r['year']}年{r['month']:02d}月"
        print(f"║  {ticker:<5} {COMPANY_NAMES.get(ticker,ticker):<8} {period:<10} {rev_b:>12.3f}億 {mom_s:>8} {yoy_s:>10}  ║")
    print("╚" + "═" * 76 + "╝")


# ── 主程式 ────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="月營收資料匯出工具")
    parser.add_argument("--format", choices=["csv","excel","both"], default="both",
                        help="輸出格式（預設 both）")
    parser.add_argument("--summary", action="store_true",
                        help="只印出最新快照，不匯出檔案")
    args = parser.parse_args()

    rows = get_all_data()
    if not rows:
        print("❌ 資料庫無數據，請先執行 fetch_revenue.py")
        return

    quarterly = compute_quarterly(rows)

    if args.summary:
        print_snapshot(rows)
        return

    print(f"\n📊 共 {len(rows)} 筆月營收記錄 | {len(quarterly)} 個完整季度\n")

    if args.format in ("csv", "both"):
        export_csv(rows, quarterly)
    if args.format in ("excel", "both"):
        export_excel(rows, quarterly)

    print_snapshot(rows)
    print(f"\n📁 輸出目錄：{REVENUE_DIR}\n")


if __name__ == "__main__":
    main()
